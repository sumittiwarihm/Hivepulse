from django.shortcuts import render, redirect
from django.http import JsonResponse, HttpResponse,response
from openpyxl import Workbook
from datetime import datetime
from openpyxl.styles import Protection
from .forms import ExcelUploadForm
from .models import amazonProduct, flipkartProduct, playstoreProduct,sentimentResult,review
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from rest_framework.permissions import IsAuthenticated
from rest_framework_simplejwt.authentication import JWTAuthentication
from django.contrib.auth.models import User
from django.contrib.auth.hashers import make_password
from django.views.decorators.csrf import csrf_exempt
from rest_framework_simplejwt.tokens import RefreshToken
import json
from django.contrib.auth import authenticate
from rest_framework.views import APIView
import openpyxl
from django.contrib.auth import logout
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework_simplejwt.authentication import JWTAuthentication
from rest_framework_simplejwt.exceptions import InvalidToken, TokenError
from django.core.management import call_command
from django.core.mail import send_mail
from django.conf import settings

#use registartion code--------------------------------------------------------------------------------------------


# user registration end------------------------------------------------------------------------------------------------------------------

# user login -------------------------------------------------------------------------------------------
from django.utils import timezone
from django.contrib.auth.models import User
@csrf_exempt
def login(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        username = data.get('username')
        password = data.get('password')

        user = authenticate(username=username, password=password)
        if user is not None:
            refresh = RefreshToken.for_user(user)
            user.last_login = timezone.now()
            user.save(update_fields=['last_login'])
            return JsonResponse({
                'refresh': str(refresh),
                'access': str(refresh.access_token),
            })
        else:
            return JsonResponse({'error': 'Invalid credentials'}, status=400)
    elif request.method == 'GET':
        # You can choose to return a JsonResponse or render a login template
        return JsonResponse({'error': 'GET method is not allowed. Please use POST method to login.'}, status=405)

    # Handle any other HTTP method if needed
    return JsonResponse({'error': 'Method not allowed'}, status=405)

    
# user login end---------------------------------------------------------------------------------------------

# @csrf_exempt
def logout_view(request):
    if request.method == 'POST':
        logout(request)
        return JsonResponse({'message': 'Logged out successfully'})
    return JsonResponse({'error': 'Invalid method'}, status=405)

# common code-------------------------------------------------------------------------------


  # this is function way of writing django rst vie funcional   
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def verify_token(request):
    # If the request reaches this point, the token is valid
    return Response({'valid': True})

#common code end--------------------------------------------------------------------------------


# redirect code start-------------------------------------------------------
def session_input_view(request):
    return render(request, 'platforms/graph.html')
def GetRawPage(request):
    return render(request,'platforms/product_sentiment.html')

def loginPage(request):
    return render(request, 'platforms/login.html')

def home(request):
    return render (request,'platforms/home.html')

def amazonPage(request):
    return render (request, 'platforms/amazonForm.html')

def flipkartPage(request):
    return render (request, 'platforms/flipkartForm.html')

def playstorePage(request):
    return render (request, 'platforms/playstoreForm.html')




# redirect code end -------------------------------------------------------------------------------------



#  amazon start----------------------------------------------------------------------------------------
class DownloadAmazonExcelTemplateView(APIView):
    # permission_classes = [IsAuthenticated]
    def get(self,request):
    # Create a new workbook and select the active worksheet
        workbook = Workbook()
        worksheet = workbook.active
        # Define the column names
        column_names = ['Asin', 'Brand']
        # Write the column names to the first row 
        for column_index, column_name in enumerate(column_names, start=1):
            worksheet.cell(row=1, column=column_index, value=column_name)
        # Create the HttpResponse object with the appropriate headers
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=excel_template_amazon.xlsx'

        # Save the workbook to the response
        workbook.save(response)

        return response
    


class uploadAmazon(APIView):
    # authentication_classes = [JWTAuthentication]
    # permission_classes = [IsAuthenticated]  these are alter nate because we alredy config thisin setting 
    def post(self,request):
        expected_columns = ['Asin', 'Brand']  # Define the expected column names
        if request.method == 'POST':
            form = ExcelUploadForm(request.POST, request.FILES)
            if form.is_valid():
                f = request.FILES['file']
                wb = openpyxl.load_workbook(f)
                ws = wb.active
                actual_columns = [cell.value for cell in ws[1]]
                if actual_columns != expected_columns:
                    return JsonResponse({'success': False, 'error': 'This file does not belong to this database downlaod Template and refill the data once again. Expected columns: ' + ', '.join(expected_columns)})
                data_rows = list(ws.iter_rows(min_row=2, values_only=True))
                
                if not data_rows:
                    return JsonResponse({'success': False, 'error': 'No data filled in your Excel file.'})
                # Get username of the logged-in user
                username = request.user.username
                # Generate session ID as a string representation of the current time
                session_id = datetime.now().strftime('%Y%m%d%H%M%S')+"_amazon"  # e.g., '20230829103015'
                
                for row in ws.iter_rows(min_row=2, values_only=True):
                    Asin, Brand = row
                    if Asin and Brand:
                        amazonProduct.objects.create(
                            Asin=Asin,
                            Brand=Brand,
                            user=username,  # Set the user
                            sessionId=session_id  # Set the session ID
                        )
                      # Send email with session ID
                subject = 'Amazon Product Upload Session ID'
                message = f'Dear {username},\n\nYour session ID for the recent Amazon product upload is: {session_id}\n\nRegards,\nYour Team'
                recipient_list = ['tiwarisumit272181@gmail.com','harishkumar.c@hiveminds.in']
                send_mail(subject, message, settings.EMAIL_HOST_USER, recipient_list, fail_silently=False)
                return JsonResponse({'success': True,'sessionId':session_id})
            else:
                return JsonResponse({'success': False, 'error': 'Form is not valid'})
        else:
            form = ExcelUploadForm()
        return render(request, 'amazonForm.html', {'form': form})
    
# logic for getting review from amazon using the asin start------------------------------------------------------------------
 # Allows POST requests without CSRF token, use with caution
# views.py
import json
from django.http import JsonResponse
from rest_framework.views import APIView
from platforms.utilsAmazonScrapping import fetch_amazon_reviews
class runAmazonReviewScrappingScript(APIView):
    def post(self, request):
        data = json.loads(request.body)
        sessionId = data.get('sessionId')
        username = request.user.username
        try:
            message = fetch_amazon_reviews(sessionId=sessionId, username=username)
            return JsonResponse({'status': 'success', 'message': message})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})


## logic for getting review from amazon using the asin end---------------------------------------------------------------



# logic for getting sentiment of those reviews and saving to data running mechanism start ------------------------------------------------------
# views.py

from .utilsAmazonSentiment import perform_amazon_sentiment_analysis

class runAmazonReviewSentimentScript(APIView):
    def post(self, request):
        data = json.loads(request.body)
        sessionId = data.get('sessionId')
        username = request.user.username

        try:
            perform_amazon_sentiment_analysis(sessionId=sessionId, username=username)
            return JsonResponse({'status': 'success', 'message': 'Amazon sentiment analysis script executed successfully.'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})


    
# logic for getting sentiment of those reviews and saving to data running mechanism end----------------------------


#  amazon end ---------------------------------------------------------------------------------------------



#  flipkart start-------------------------------------------------------------------------------------------



class downloadFlipkartExcelTemplate(APIView):
    # permission_classes = [IsAuthenticated]
    def get(self,request):
 # Create a new workbook and select the active worksheet
        workbook = Workbook()
        worksheet = workbook.active
        # Define the column names
        column_names = ['Fsn', 'Brand']
        # Write the column names to the first row 
        for column_index, column_name in enumerate(column_names, start=1):
            cell = worksheet.cell(row=1, column=column_index, value=column_name)
        # Create the HttpResponse object with the appropriate headers
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=excel_template_flipkart.xlsx'

        # Save the workbook to the response
        workbook.save(response)

        return response



class uploadFlipkart(APIView):
    # authentication_classes = [JWTAuthentication]
    # permission_classes = [IsAuthenticated]
    def post(self , request):
        expected_columns = ['Fsn', 'Brand']  # Define the expected column names
        if request.method == 'POST':
            form = ExcelUploadForm(request.POST, request.FILES)
            if form.is_valid():
                f = request.FILES['file']
                wb = openpyxl.load_workbook(f)
                ws = wb.active
                actual_columns = [cell.value for cell in ws[1]]
                if actual_columns != expected_columns:
                    return JsonResponse({'success': False, 'error': 'This file does not belong to this database. Expected columns: ' + ', '.join(expected_columns)})
                data_rows = list(ws.iter_rows(min_row=2, values_only=True))
                if not data_rows:
                    return JsonResponse({'success': False, 'error': 'No data filled in your Excel file.'})
                username=request.user.username
                session_id=datetime.now().strftime('%Y%m%d%H%M%S')+"_flipkart"
                for row in ws.iter_rows(min_row=2, values_only=True):
                    Fsn, Brand = row
                    if Fsn and Brand:
                        flipkartProduct.objects.create(
                            Fsn=Fsn, 
                            Brand=Brand,
                            user=username,
                            sessionId=session_id
                            )
                subject="Flipkart Product Upload Session ID"
                message=f'Dear {username},\n\n Your session ID for the recent flipkart product upload is :{session_id}\n\nRegrads,\nAnalytics Team'
                recipient_list = ['tiwarisumit272181@gmail.com','harishkumar.c@hiveminds.in']
                send_mail(subject, message, settings.EMAIL_HOST_USER, recipient_list, fail_silently=False)
                return JsonResponse({'success': True,'sessionId':session_id})
            else:
                return JsonResponse({'success': False, 'error': 'Form is not valid'})
        else:
            form = ExcelUploadForm()
        return render(request, 'flipkartForm.html', {'form': form})



# views.py

from .utilsFlipkartScrapping import fetch_flipkart_reviews

class runFlipkartReviewScrappingScript(APIView):
    def post(self, request):
        data = json.loads(request.body)
        sessionId = data.get('sessionId')
        username = request.user.username

        try:
            message = fetch_flipkart_reviews(sessionId=sessionId, username=username)
            return JsonResponse({'status': 'success', 'message': message})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})



# views.py

from .utilsFlipkartSentiment import perform_flipkart_sentiment_analysis

class runFlipkartReviewSentimentScript(APIView):
    def post(self, request):
        if request.method == 'POST':
            data = json.loads(request.body)
            sessionId = data.get('sessionId')
            username = request.user.username

            try:
                message = perform_flipkart_sentiment_analysis(sessionId=sessionId, username=username)
                return JsonResponse({'status': 'success', 'message': message})
            except Exception as e:
                return JsonResponse({'status': 'error', 'message': str(e)})
        return JsonResponse({'status': 'error', 'message': 'Invalid request method.'})


# flipkart end ----------------------------------------------------------------------------------------------



#  playstore start------------------------------------------------------------------------------------------


class downloadPlaystoreExcelTemplate(APIView):
    # permission_classes = [IsAuthenticated]
    def get(self,request):

    # Create a new workbook and select the active worksheet
        workbook = Workbook()
        worksheet = workbook.active
        # Define the column names
        column_names = ['AppId', 'Brand']
        # Write the column names to the first row 
        for column_index, column_name in enumerate(column_names, start=1):
            cell = worksheet.cell(row=1, column=column_index, value=column_name)
        # Create the HttpResponse object with the appropriate headers
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=excel_template_playstore.xlsx'

        # Save the workbook to the response
        workbook.save(response)

        return response


class uploadPlaystore(APIView):
    # authentication_classes = [JWTAuthentication]
    # permission_classes = [IsAuthenticated]
    def post(self , request):
        expected_columns = ['AppId', 'Brand']  # Define the expected column names
        if request.method == 'POST':
            form = ExcelUploadForm(request.POST, request.FILES)
            if form.is_valid():
                f = request.FILES['file']
                wb = openpyxl.load_workbook(f)
                ws = wb.active
                actual_columns = [cell.value for cell in ws[1]]
                if actual_columns != expected_columns:
                    return JsonResponse({'success': False, 'error': 'This file does not belong to this database. Expected columns: ' + ', '.join(expected_columns)})
                data_rows = list(ws.iter_rows(min_row=2, values_only=True))
                if not data_rows:
                    return JsonResponse({'success': False, 'error': 'No data filled in your Excel file.'})
                username = request.user.username
                # Generate session ID as a string representation of the current time
                session_id = datetime.now().strftime('%Y%m%d%H%M%S')+"_playstore"
                for row in ws.iter_rows(min_row=2, values_only=True):
                    AppId, Brand = row
                    playstoreProduct.objects.create(
                        AppId=AppId,
                        Brand=Brand,
                        user=username,  # Set the user
                        sessionId=session_id 
                        )
                try:
                    subject = 'Playstore Product Upload Session ID'
                    message = f'Dear {username},\n\nYour session ID for the recent Playstore product upload is: {session_id}\n\nRegards,\nYour Team'
                    recipient_list = ['tiwarisumit272181@gmail.com','harishkumar.c@hiveminds.in']
                    send_mail(subject, message, settings.EMAIL_HOST_USER, recipient_list, fail_silently=False)
                    return JsonResponse({'success': True,'sessionId':session_id})
                except:
                    return JsonResponse({'success':False,'error':'your email is not responding'})
            else:
                return JsonResponse({'success': False, 'error': 'Form is not valid'})
        else:
            form = ExcelUploadForm()
        return render(request, 'playstoreForm.html', {'form': form})
    

# views.py

from .utilsPlaystoreScrapping import fetch_playstore_reviews

class runPlaystoreReviewScrappingScript(APIView):
    def post(self, request):
        if request.method == 'POST':
            data = json.loads(request.body)
            sessionId = data.get('sessionId')
            username = request.user.username

            try:
                message = fetch_playstore_reviews(sessionId=sessionId, username=username)
                return JsonResponse({'status': 'success', 'message': message})
            except Exception as e:
                return JsonResponse({'status': 'error', 'message': str(e)})
        return JsonResponse({'status': 'error', 'message': 'Invalid request method.'})




# views.py

from .utilsPlaystoreSentiment import perform_playstore_sentiment_analysis

class runPlaystoreReviewSentimentScript(APIView):
    def post(self, request):
        if request.method == 'POST':
            data = json.loads(request.body)
            sessionId = data.get('sessionId')
            username = request.user.username

            try:
                message = perform_playstore_sentiment_analysis(sessionId=sessionId, username=username)
                return JsonResponse({'status': 'success', 'message': message})
            except Exception as e:
                return JsonResponse({'status': 'error', 'message': str(e)})
        return JsonResponse({'status': 'error', 'message': 'Invalid request method.'})

    # playstore end --------------------------------------------------------------------------------------------------------

# cross platform redirecting----------------------------------------------------------------------------------------------------------------

# from django.shortcuts import render
# from django.contrib.contenttypes.models import ContentType
# from platforms.models import amazonProduct, flipkartProduct, playstoreProduct, review
# # @ login_required
# def product_sentiment_view(request):
#     results = []
#     error = None
#     platform = request.POST.get('platform', '')
#     session_id = request.POST.get('sessionId', '')

#     if request.method == 'POST':
#         platform = request.POST.get('platform')
#         identifier = request.POST.get('sessionId')
#         platform_mapping = {
#             'amazon': (amazonProduct, 'Asin'),
#             'flipkart': (flipkartProduct, 'Fsn'),
#             'playstore': (playstoreProduct, 'AppId'),
#         }
        
#         if platform in platform_mapping:
#             model_class, id_field = platform_mapping[platform]
#         else:
#             error = "Invalid platform selected."
        
#         if identifier and not error:
#             try:
#                 products = model_class.objects.filter(sessionId=identifier, Status='completed')
                
#                 if not products.exists():
#                     error = f"No completed product with the identifier {identifier} found in {platform.capitalize()}."
#                 else:
#                     for product in products:
#                         product_id = getattr(product, id_field)
#                         product_Brand = getattr(product, 'Brand')
                        
#                         content_type = ContentType.objects.get_for_model(model_class)

#                         reviews_queryset = review.objects.filter(
#                             content_type=content_type, 
#                             object_id=product.id
#                         )
                        
#                         for rev in reviews_queryset:
#                             sentiment_results = rev.sentimentresult_set.all()
#                             for sentiment in sentiment_results:
#                                 results.append((
#                                     product_id,
#                                     product_Brand,
#                                     rev.reviewContent, 
#                                     sentiment.estimatedResult,
#                                     rev.rating, 
#                                     round(sentiment.positiveScore, 3), 
#                                     round(sentiment.neutralScore, 3), 
#                                     round(sentiment.negativeScore, 3),
#                                 ))

#                 if not results:
#                     error = "No comments or sentiment results found for the provided identifier."
                
#             except model_class.DoesNotExist:
#                 error = f"Product with the provided identifier does not exist in the {platform.capitalize()} platform or is not yet completed."
                

#     return render(request, 'platforms/product_sentiment.html', {
#         'results': results,
#         'error': error,
#         'platform': platform,
#         'sessionId': session_id
#     })




from django.shortcuts import render
from django.http import JsonResponse
from django.contrib.contenttypes.models import ContentType
from platforms.models import amazonProduct, flipkartProduct, playstoreProduct, review

# @login_required
@csrf_exempt
def product_sentiment_view(request):
    if request.method != 'POST':
        return JsonResponse({'error': "Method not allowed. Only POST requests are accepted."}, status=405)
    results = []
    body = json.loads(request.body)
    platform = body.get('platform',' ' )
    session_id = body.get('sessionId', ' ')
    if not platform:
        return JsonResponse({'error': "Platform is required."}, status=400)
    if not session_id:
        return JsonResponse({'error': "Session ID is required."}, status=400)

    platform_mapping = {
        'amazon': (amazonProduct, 'Asin'),
        'flipkart': (flipkartProduct, 'Fsn'),
        'playstore': (playstoreProduct, 'AppId'),
    }

    if platform not in platform_mapping:
        return JsonResponse({'error': "Invalid platform selected. Choose from 'amazon', 'flipkart', or 'playstore'."}, status=400)

    model_class, id_field = platform_mapping[platform]

    try:
        # Check if any completed products exist for the given session ID
        products = model_class.objects.filter(sessionId=session_id, Status='completed')

        if not products.exists():
            return JsonResponse({'error': f"No completed products found for the provided session ID: {session_id}."}, status=404)

        # Process each product
        for product in products:
            product_id = getattr(product, id_field, None)
            product_brand = getattr(product, 'Brand', None)

            if not product_id or not product_brand:
                continue  # Skip products with missing critical information

            # Fetch reviews associated with the product
            content_type = ContentType.objects.get_for_model(model_class)
            reviews_queryset = review.objects.filter(content_type=content_type, object_id=product.id)

            for rev in reviews_queryset:
                sentiment_results = rev.sentimentresult_set.all()
                for sentiment in sentiment_results:
                    results.append({
                        'product_id': product_id,
                        'brand': product_brand,
                        'review_content': rev.reviewContent,
                        'sentiment_result': sentiment.estimatedResult,
                        'rating': rev.rating,
                        'positive_score': round(sentiment.positiveScore, 3),
                        'neutral_score': round(sentiment.neutralScore, 3),
                        'negative_score': round(sentiment.negativeScore, 3),
                    })

        if not results:
            return JsonResponse({'error': f"No sentiment results found for session ID: {session_id}."}, status=404)

    except model_class.DoesNotExist:
        return JsonResponse({'error': f"No records found for the session ID: {session_id} in the {platform.capitalize()} platform."}, status=404)
    except Exception as e:
        return JsonResponse({'error': f"An unexpected error occurred: {str(e)}"}, status=500)

    return JsonResponse({'data': results}, status=200)

#----------------------------------------------------------------------------------------------
import pandas as pd
from django.http import HttpResponse
from django.contrib.contenttypes.models import ContentType
from platforms.models import amazonProduct, flipkartProduct, playstoreProduct, review
# @ login_required
def download_excel(request):
    platform = request.POST.get('platform')
    identifier = request.POST.get('sessionId')
    
    platform_mapping = {
        'amazon': (amazonProduct, 'Asin'),
        'flipkart': (flipkartProduct, 'Fsn'),
        'playstore': (playstoreProduct, 'AppId'),
    }

    model_class, id_field = platform_mapping.get(platform, (None, None))

    if model_class and identifier:
        products = model_class.objects.filter(sessionId=identifier, Status='completed')
        data = []
        
        for product in products:
            product_id = getattr(product, id_field)
            product_Brand = getattr(product, 'Brand')

            content_type = ContentType.objects.get_for_model(model_class)
            reviews_queryset = review.objects.filter(
                content_type=content_type, 
                object_id=product.id
            )

            for rev in reviews_queryset:
                sentiment_results = rev.sentimentresult_set.all()
                for sentiment in sentiment_results:
                    data.append({
                        'ID': product_id,
                        'Brand': product_Brand,
                        'Comment': rev.reviewContent,
                        'Sentiment Result': sentiment.estimatedResult,
                        'Rating': rev.rating,
                        'Positive Score': round(sentiment.positiveScore, 3),
                        'Neutral Score': round(sentiment.neutralScore, 3),
                        'Negative Score': round(sentiment.negativeScore, 3),
                    })

        if data:
            df = pd.DataFrame(data)
            response = HttpResponse(content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = 'attachment; filename="sentiment_results.xlsx"'
            df.to_excel(response, index=False)
            return response
        else:
            return HttpResponse("No data found for the provided identifier.")
    return HttpResponse("Invalid platform/session or no identifier provided.")


#---------------------------------------------------------------------------experiment-----------------------------------------------------


from collections import defaultdict
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import logging

@csrf_exempt
def sessionInputPlaystore(request):
    return render(request, 'platforms/sessionInputPlaystore.html')
def sessionInputAmazon(request):
    return render (request,'platforms/sessionInputAmazon.html')
def sessionInputFlipkart(request):
    return render (request,'platforms/sessionInputFlipkart.html')
@ login_required
def getDataForGraph(request):
    if request.method == 'GET':  # Ensure it's a POST request
        return JsonResponse({'error': 'wrong request method'}, status=400)
    body = json.loads(request.body)
    sessionId = body.get('sessionId')  # Get sessionId from POST data
    if not sessionId:
        return JsonResponse({'error': 'sessionId is required'}, status=400)

    try:
        # Query for products based on the sessionId
        amazon_product = amazonProduct.objects.filter(sessionId=sessionId)
        amazon_product_type = ContentType.objects.get_for_model(amazonProduct)
    except amazonProduct.DoesNotExist:
        return JsonResponse({'error': 'No products found for the given sessionId'}, status=404)
    except Exception as e:
        logging.error(f"Error fetching products or content type: {e}")
        return JsonResponse({'error': 'An unexpected error occurred'}, status=500)

    # Dictionary to store counts of positive, negative, and neutral reviews for each ASIN
    asin_sentiment_counts = defaultdict(lambda: {'positiveCount': 0, 'negativeCount': 0, 'neutralCount': 0})

    # Iterate through products and reviews
    try:
        for product in amazon_product:
            product_reviews = review.objects.filter(content_type=amazon_product_type, object_id=product.id)
            review_ids = [review.id for review in product_reviews]
            sentiments = sentimentResult.objects.filter(review_id__in=review_ids)

            for sentiment in sentiments:
                if sentiment.estimatedResult == 'Positive':
                    asin_sentiment_counts[product.Asin]['positiveCount'] += 1
                elif sentiment.estimatedResult == 'Negative':
                    asin_sentiment_counts[product.Asin]['negativeCount'] += 1
                elif sentiment.estimatedResult == 'Neutral':
                    asin_sentiment_counts[product.Asin]['neutralCount'] += 1
    except Exception as e:
        logging.error(f"Error processing reviews or sentiments: {e}")
        return JsonResponse({'error': 'An error occurred while processing reviews and sentiments'}, status=500)

    # Prepare data to send to frontend
    data = []
    for asin, counts in asin_sentiment_counts.items():
        data.append({
            'Asin': asin,
            'positiveCount': counts['positiveCount'],
            'negativeCount': counts['negativeCount'],
            'neutralCount': counts['neutralCount']
        })

    return JsonResponse({'data': data})





#----------------------------------------------------------------------------------------------------------

# reviews/views.py

from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import json
from .models import playstoreProduct, review, sentimentResult
from django.contrib.contenttypes.models import ContentType
from collections import defaultdict
from .utils import assign_category, assign_category2

# @ login_required
@csrf_exempt
def getDataforPlaystoreCategorization(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'Only POST method is allowed'}, status=400)
    try:
        body = json.loads(request.body)
        sessionId = body.get('sessionId')
        if not sessionId:
            return JsonResponse({'error': 'sessionId is required'}, status=400)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON payload'}, status=400)

    try:
        # Fetch all playstoreProduct instances related to the given sessionId
        playstore_products = playstoreProduct.objects.filter(sessionId=sessionId)
        # print(playstore_products)
        if not playstore_products.exists():
            return JsonResponse({'error': 'No apps found for the provided sessionId.'}, status=404)
        
        # Prepare response data
        response_data = []

        # For each AppId, get the corresponding reviews and sentiments
        for product in playstore_products:
            appId = product.AppId
            total_positive = total_negative = total_neutral = 0

            # Initialize sentiment counts per category
            category_sentiments = defaultdict(lambda: {'positive': 0, 'negative': 0, 'neutral': 0})

            # Get the content type for playstoreProduct
            product_type = ContentType.objects.get_for_model(playstoreProduct)

            # Fetch all reviews for this AppId
            reviews = review.objects.filter(content_type=product_type, object_id=product.id)

            for rev in reviews:
                # Fetch sentiment result from sentimentResult model (already stored)
                sentiment = sentimentResult.objects.filter(review=rev.id).first()
                sentiment_result = sentiment.estimatedResult.lower() if sentiment else 'neutral'

                # Categorize the review based on its content
                category = assign_category(rev.reviewContent)

                # Update category-wise sentiment counts using the stored sentiment result
                if sentiment_result == 'positive':
                    total_positive += 1
                    category_sentiments[category]['positive'] += 1
                elif sentiment_result == 'negative':
                    total_negative += 1
                    category_sentiments[category]['negative'] += 1
                else:
                    total_neutral += 1
                    category_sentiments[category]['neutral'] += 1

            # Add aggregated data for this app to the response
            response_data.append({
                'appId': product.Brand,
                'totalPositive': total_positive,
                'totalNegative': total_negative,
                'totalNeutral': total_neutral,
                'category': category_sentiments
            })

        return JsonResponse({'data': response_data}, status=200)

    except Exception as e:
        return JsonResponse({'error': f'An error occurred: {str(e)}'}, status=500)
    
            

#---------------graph api building for amazon--------------------------------------
# i want data in formate of (asin , brand totalPositive, totalNegative ,totalNeutral)
@csrf_exempt
def getDataForAmazonCategorization(request):
    if(request.method!='POST'):
        return JsonResponse({'error': 'Only POST method is allowed'}, status=400)
    try:
       body=json.loads(request.body)
       sessionId=body.get('sessionId') 
       if not sessionId:
           return  JsonResponse({'error':"sessionId is required" },status=400)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON payload'}, status=400)
    try:
        amazon_products=amazonProduct.objects.filter(sessionId=sessionId)
        if not amazon_products.exists():
            return JsonResponse({'error':"no amazon ASIN is present in tha databse"})
        response_data=[]
        for product in amazon_products:
            Asin=product.Asin
            category_sentiments = defaultdict(lambda: {'positive': 0, 'negative': 0, 'neutral': 0})
            total_positive=total_negative=total_neutral=0
            product_type=ContentType.objects.get_for_model(amazonProduct)
            reviews=review.objects.filter(content_type=product_type,object_id=product.id)
            for rev in reviews:
                sentiment=sentimentResult.objects.filter(review_id=rev.id).first()
                sentiment_result=sentiment.estimatedResult.lower() if sentiment else 'neutral'
                category = assign_category2(rev.reviewContent)
                # Update category-wise sentiment counts using the stored sentiment result
                if sentiment_result == 'positive':
                    total_positive += 1
                    category_sentiments[category]['positive'] += 1
                elif sentiment_result == 'negative':
                    total_negative += 1
                    category_sentiments[category]['negative'] += 1
                else:
                    total_neutral += 1
                    category_sentiments[category]['neutral'] += 1
            response_data.append({
                'asin': Asin,
                'brand': product.Brand,
                'totalPositive':total_positive,
                'totalNegative':total_negative,
                'totalNeutral':total_neutral,
                'category': category_sentiments
            })
        return JsonResponse({'data':response_data},status=200)
    except Exception as e:
        return JsonResponse({'error':f"an error occured during interaction with database :{str(e)}"},status=500)
# seeee below this




















































@csrf_exempt
def getDataForFlipkartCategorization(request):
    if(request.method!='POST'):
        return JsonResponse({'error':"only Post method is alllowed"},status=400)
    try:
        body=json.loads(request.body)
        sessionId=body.get('sessionId')
        if not sessionId:
            return JsonResponse({'error':'sessionID is not found'},status=400)
    except json.JSONDecodeError:
        return JsonResponse({'error':"invalid payload"},status=400)
    try:
        flipkart_products=flipkartProduct.objects.filter(sessionId=sessionId)
        if not flipkart_products.exists():
            return JsonResponse({'error':"no product exist for this sessionId"})
        response_data=[]
        try:
            for product in flipkart_products:
                category_sentiments = defaultdict(lambda: {'positive': 0, 'negative': 0, 'neutral': 0})
                # now for every product we want all the reviews and after that we will require all the sentiment according to that
                total_positive=total_negative=total_neutral=0
                Fsn=product.Fsn
                product_type=ContentType.objects.get_for_model(flipkartProduct)
                reviews=review.objects.filter(content_type=product_type,object_id=product.id)
                for rev in reviews:
                    sentiment=sentimentResult.objects.filter(review_id=rev.id).first()
                    sentiment_result=sentiment.estimatedResult.lower() if sentiment else 'neutral'
                    category = assign_category2(rev.reviewContent)
                    if sentiment_result == 'positive':
                        total_positive += 1
                        category_sentiments[category]['positive'] += 1
                    elif sentiment_result == 'negative':
                        total_negative += 1
                        category_sentiments[category]['negative'] += 1
                    else:
                        total_neutral += 1
                        category_sentiments[category]['neutral'] += 1
                try: 
                    response_data.append({
                        'fsn':Fsn,
                        'brand':product.Brand,
                        'totalPositive':total_positive,
                        'totalNegative':total_negative,
                        'totalNeutral': total_neutral,
                        'category': category_sentiments
                    })
                except Exception as e:
                    return JsonResponse({'error':"error in interacting with database"},status=404)
            return JsonResponse({'data': response_data}, status=200)
        except:
            return JsonResponse({'error':"error in going through product "},status=404)
    except Exception as e:
        return JsonResponse({'error':f"an error occured during interaction with database :{str(e)}"},status=500)

                

# def getDataForWorldCloudForAmazon(request):
#     if(request.method!='POST'):
#         return JsonResponse({'error':"only post method is allowed"},status=400)
#     try:
#         body=json.loads(request.body)
#         sessionId=body.get('sessionId')
#         if not sessionId:
#             return JsonResponse({'error':"session id is not found or payload is invalid"},status=400)
#     except json.JSONDecodeError:
#         return JsonResponse({'error':"invalid payload"},status=400)
#     try:
#         reviews=review.objects.filter(sessionId=sessionId)
#         if not review.exists():
#             return JsonResponse({'error':"no review exist for this sessionId"})





             











        

        








